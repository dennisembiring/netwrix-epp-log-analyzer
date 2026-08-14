"""Export helpers for CSV / Excel / PDF."""

import io
from xml.sax.saxutils import escape

import pandas as pd
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


def to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8")


def to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Data") -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        for i, col_name in enumerate(df.columns, start=1):
            max_len = max(
                [len(str(col_name))] + [len(str(v)) for v in df[col_name].head(200)]
            )
            ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 60)
    return buf.getvalue()


def to_full_report_excel_bytes(sections: list[tuple[str, pd.DataFrame]]) -> bytes:
    """One workbook, one sheet per labeled section, in filed order."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        used_names = set()
        for label, df in sections:
            if df is None or df.empty:
                continue
            # Excel sheet names: max 31 chars, no []:*?/\\, must be unique.
            name = "".join(c for c in label if c not in '[]:*?/\\')[:31] or "Sheet"
            base, n = name, 1
            while name in used_names:
                suffix = f" ({n})"
                name = base[: 31 - len(suffix)] + suffix
                n += 1
            used_names.add(name)

            df.to_excel(writer, index=False, sheet_name=name)
            ws = writer.sheets[name]
            for i, col_name in enumerate(df.columns, start=1):
                max_len = max(
                    [len(str(col_name))] + [len(str(v)) for v in df[col_name].head(200)]
                )
                ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 60)
    return buf.getvalue()


_CELL_STYLE = ParagraphStyle(
    name="cell", fontSize=7, leading=9, wordWrap="CJK"
)
_HEADER_CELL_STYLE = ParagraphStyle(
    name="cell_header", fontSize=7, leading=9, textColor=colors.white, wordWrap="CJK"
)


# A landscape letter page has ~10in of usable width; below this floor,
# text becomes illegible regardless of the wrap/width math, and above
# roughly 12-15 columns the per-column share drops below that floor no
# matter how it's distributed. Cap columns instead of trying to cram
# arbitrarily many into a fixed-width page.
_MAX_PDF_COLUMNS = 12
_MIN_COL_WIDTH = 0.6 * inch


def _cap_columns(df: pd.DataFrame, max_cols: int = _MAX_PDF_COLUMNS) -> tuple[pd.DataFrame, int]:
    """Returns (possibly-truncated df, total original column count)."""
    if len(df.columns) <= max_cols:
        return df, len(df.columns)
    return df.iloc[:, :max_cols], len(df.columns)


def _data_table(df: pd.DataFrame, max_rows: int, avail_width: float) -> Table:
    shown = df.head(max_rows)
    header = [Paragraph(escape(str(c)), _HEADER_CELL_STYLE) for c in shown.columns]
    body = [
        [Paragraph(escape(str(v)), _CELL_STYLE) for v in row]
        for row in shown.astype(str).values.tolist()
    ]
    data = [header] + body

    # Give wide-content columns (long URLs/domains) more room than short
    # categorical ones instead of splitting the page evenly, so text wraps
    # into a few readable lines rather than one over-squeezed column.
    sample = shown.astype(str)
    raw_widths = [
        max([len(str(c))] + [len(v) for v in sample[c].head(50)]) for c in shown.columns
    ]
    n = len(raw_widths)
    if n * _MIN_COL_WIDTH >= avail_width:
        # Even the floor doesn't fit (shouldn't happen once callers cap
        # column count, but stay safe) -- split evenly rather than risk a
        # negative width.
        col_widths = [avail_width / n] * n
    else:
        # Columns below the proportional floor get exactly the floor; the
        # rest of the width is distributed proportionally among the
        # remaining columns only, so a later rescale can never push anyone
        # back under the floor (the bug that crashed on wide tables).
        total = sum(raw_widths) or 1
        floored = [avail_width * w / total < _MIN_COL_WIDTH for w in raw_widths]
        reserved = sum(_MIN_COL_WIDTH for f in floored if f)
        remaining_width = avail_width - reserved
        remaining_total = sum(w for w, f in zip(raw_widths, floored) if not f) or 1
        col_widths = [
            _MIN_COL_WIDTH if f else remaining_width * w / remaining_total
            for w, f in zip(raw_widths, floored)
        ]

    table = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2b3a55")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
            ]
        )
    )
    return table


def _needs_landscape(df: pd.DataFrame | None) -> bool:
    """Landscape pays off for wide tables: many columns, or columns with long
    text content (URLs/domains in Matched Item, Destination Details, etc.)
    that would otherwise get squeezed illegibly in portrait."""
    if df is None or df.empty:
        return False
    if len(df.columns) > 4:
        return True
    sample = df.astype(str)
    return any(sample[c].head(50).str.len().max() > 25 for c in sample.columns)


def to_pdf_bytes(
    title: str,
    summary_lines: list[str],
    df: pd.DataFrame | None = None,
    max_rows: int = 200,
) -> bytes:
    buf = io.BytesIO()
    # Landscape only pays off for wide tables; narrow summary tables (like the
    # Reports page breakdowns) look lost in all that empty space otherwise.
    page_size = landscape(letter) if _needs_landscape(df) else letter
    margin = 0.5 * inch
    doc = SimpleDocTemplate(
        buf,
        pagesize=page_size,
        topMargin=margin,
        bottomMargin=margin,
        leftMargin=margin,
        rightMargin=margin,
    )
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    title_style.alignment = TA_LEFT
    story = [Paragraph(title, title_style), Spacer(1, 12)]

    for line in summary_lines:
        story.append(Paragraph(line, styles["Normal"]))
    story.append(Spacer(1, 16))

    if df is not None and not df.empty:
        avail_width = page_size[0] - 2 * margin
        shown_df, orig_cols = _cap_columns(df)
        story.append(_data_table(shown_df, max_rows, avail_width))
        notes = []
        if len(df) > max_rows:
            notes.append(f"Showing {max_rows} of {len(df)} rows.")
        if orig_cols > len(shown_df.columns):
            notes.append(
                f"Showing the first {len(shown_df.columns)} of {orig_cols} columns."
            )
        if notes:
            notes.append("Use CSV/Excel export for the complete data.")
            story.append(Spacer(1, 8))
            story.append(Paragraph(" ".join(notes), styles["Normal"]))

    doc.build(story)
    return buf.getvalue()


def to_full_report_pdf_bytes(
    title: str,
    summary_lines: list[str],
    sections: list[tuple[str, pd.DataFrame]],
    max_rows: int = 100,
) -> bytes:
    """One PDF combining several labeled summary tables, in filed order."""
    buf = io.BytesIO()
    sections = [(label, df) for label, df in sections if df is not None and not df.empty]
    # Landscape once for the whole document if any section needs the extra
    # width, rather than per-section, since ReportLab renders one continuous
    # flow and can't flip orientation mid-document.
    page_size = landscape(letter) if any(_needs_landscape(df) for _, df in sections) else letter
    margin = 0.5 * inch
    doc = SimpleDocTemplate(
        buf,
        pagesize=page_size,
        topMargin=margin,
        bottomMargin=margin,
        leftMargin=margin,
        rightMargin=margin,
    )
    avail_width = page_size[0] - 2 * margin
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    title_style.alignment = TA_LEFT
    heading_style = styles["Heading2"]
    heading_style.alignment = TA_LEFT

    story = [Paragraph(title, title_style), Spacer(1, 12)]
    for line in summary_lines:
        story.append(Paragraph(line, styles["Normal"]))
    story.append(Spacer(1, 16))

    for i, (label, df) in enumerate(sections):
        story.append(Paragraph(label, heading_style))
        story.append(Spacer(1, 6))
        shown_df, orig_cols = _cap_columns(df)
        story.append(_data_table(shown_df, max_rows, avail_width))
        notes = []
        if len(df) > max_rows:
            notes.append(f"Showing {max_rows} of {len(df)} rows.")
        if orig_cols > len(shown_df.columns):
            notes.append(
                f"Showing the first {len(shown_df.columns)} of {orig_cols} columns."
            )
        if notes:
            notes.append("Use Excel/CSV export for the complete data.")
            story.append(Spacer(1, 4))
            story.append(Paragraph(" ".join(notes), styles["Normal"]))
        if i < len(sections) - 1:
            story.append(Spacer(1, 20))

    doc.build(story)
    return buf.getvalue()
